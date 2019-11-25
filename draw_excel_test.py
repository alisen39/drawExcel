import random

from PIL import Image
import openpyxl
from openpyxl.styles import fills
import os




def int_to_16(num):
    num1 = hex(num).replace('0x', '')
    num2 = num1 if len(num1) > 1 else '0' + num1
    return num2


def draw_color():
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    count = 0
    MAX_WIDTH = 255
    for w in range(1, MAX_WIDTH + 1):

        for h in range(1, MAX_WIDTH + 1):

            cell = worksheet.cell(column=w, row=h)

            if h == 1:
                _w = cell.column
                _h = cell.col_idx
                # 调整列宽
                worksheet.column_dimensions[_w].width = 1

            # 调整行高
            worksheet.row_dimensions[h].height = 6

            if count < 255 ** 3:
                back = int_to_16(num=count)
                back = '0' * (6 - len(back)) + back
            else:
                back = ''.join([int_to_16(random.randint(0, 255)) for _ in range(3)])

            cell.fill = fills.PatternFill(fill_type="solid", fgColor=back)
            count += 1

        print('write in:', w, '  |  all:', w + 1)
    print('saving...')
    workbook.save('test.xlsx')
    print('success!')


def draw_cell():
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    count = 0
    max_cell = 10000
    for w in range(1, max_cell + 1):

        for h in range(1, max_cell + 1):

            cell = worksheet.cell(column=w, row=h)

            if h == 1:
                _w = cell.column
                _h = cell.col_idx
                # 调整列宽
                worksheet.column_dimensions[_w].width = 1

            # 调整行高
            worksheet.row_dimensions[h].height = 6

            # if count < 255 ** 3:
            #     back = int_to_16(num=count)
            #     back = '0' * (6 - len(back)) + back
            # else:
            #     back = ''.join([int_to_16(random.randint(0, 255)) for _ in range(3)])

            cell.fill = fills.PatternFill(fill_type="solid", fgColor='333333')
            count += 1

        print('write in:', w, '  |  all:', w + 1)
    print('saving...')
    workbook.save('test.xlsx')
    print('success!')


if __name__ == '__main__':
    draw_color()

    # draw_cell()
